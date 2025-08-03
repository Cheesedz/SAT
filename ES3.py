import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
import sys
from pysat.solvers import Glucose3
import time
from threading import Timer, Thread, Event
import os
import ast

sat_solver = Glucose3
time_budget = 1200  # Thời gian tối đa (giây)
type = "es3"
id_counter = 1

# Mở file log ở chế độ append
log_file = open('run.log', 'a')

def write_to_xlsx(result_dict):
    excel_results = [result_dict]
    output_path = 'out/'
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)
    else:
        df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print_to_console_and_log(f"Kết quả đã được thêm vào file Excel: {os.path.abspath(excel_file_path)}\n")

def print_to_console_and_log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=log_file, **kwargs)
    log_file.flush()

def encode_problem_es3(tasks, resources):
    max_time = max(task[2] for task in tasks)
    n_tasks = len(tasks)

    # Biến u[i][j]: Tác vụ i được gán cho tài nguyên j
    u = [[i * resources + j + 1 for j in range(resources)] for i in range(n_tasks)]

    # Biến z[i][t]: Tác vụ i đang truy cập tài nguyên tại thời điểm t
    z = [[n_tasks * resources + i * max_time + t + 1 for t in range(tasks[i][0], tasks[i][2])] 
         for i in range(n_tasks)]

    # Biến D[i][j][t]: Tác vụ i bắt đầu truy cập tài nguyên j tại thời điểm t
    D = [[[n_tasks * (resources + max_time) + i * resources * max_time + j * max_time + t + 1 
          for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1)] 
         for j in range(resources)] for i in range(n_tasks)]

    # Ràng buộc D1: Một tác vụ không được gán cho hai tài nguyên cùng lúc
    for i in range(n_tasks):
        for j in range(resources):
            for jp in range(j + 1, resources):
                sat_solver.add_clause([-u[i][j], -u[i][jp]])

    # Ràng buộc D2: Mỗi tác vụ phải được gán cho ít nhất một tài nguyên
    for i in range(n_tasks):
        sat_solver.add_clause([u[i][j] for j in range(resources)])

    # Ràng buộc D3: Một tài nguyên chỉ được giữ bởi tối đa một tác vụ tại một thời điểm
    for i in range(n_tasks):
        for ip in range(i + 1, n_tasks):
            for j in range(resources):
                for t in range(max(tasks[i][0], tasks[ip][0]), min(tasks[i][2], tasks[ip][2])):
                    t_idx_i = t - tasks[i][0]
                    t_idx_ip = t - tasks[ip][0]
                    sat_solver.add_clause([-z[i][t_idx_i], -u[i][j], -z[ip][t_idx_ip], -u[ip][j]])

    # Ràng buộc D4: Mỗi tác vụ phải có chính xác một thời điểm bắt đầu
    for i in range(n_tasks):
        clause = []
        for j in range(resources):
            for t in range(tasks[i][0], tasks[i][2] - tasks[i][1] + 1):
                t_idx = t - tasks[i][0]
                clause.append(D[i][j][t_idx])
        sat_solver.add_clause(clause)

    # Ràng buộc D5: Liên kết biến bắt đầu với biến trạng thái và gán tài nguyên
    for i in range(n_tasks):
        r_i, e_i, d_i = tasks[i]
        for j in range(resources):
            for t in range(r_i, d_i - e_i + 1):
                t_idx = t - r_i
                # Nếu D[i][j][t] đúng, thì u[i][j] phải đúng
                sat_solver.add_clause([-D[i][j][t_idx], u[i][j]])
                # Nếu D[i][j][t] đúng, z[i][t'] đúng trong suốt e_i
                for tp in range(t, t + e_i):
                    tp_idx = tp - r_i
                    sat_solver.add_clause([-D[i][j][t_idx], z[i][tp_idx]])
                # Nếu D[i][j][t] đúng, z[i][t'] sai trước t
                for tp in range(r_i, t):
                    tp_idx = tp - r_i
                    sat_solver.add_clause([-D[i][j][t_idx], -z[i][tp_idx]])
                # Nếu D[i][j][t] đúng, z[i][t'] sai sau t + e_i - 1
                for tp in range(t + e_i, d_i):
                    tp_idx = tp - r_i
                    sat_solver.add_clause([-D[i][j][t_idx], -z[i][tp_idx]])

    return u, z, D

def interrupt(solver):
    solver.interrupt()

def validate_solution(tasks, model, u, z, D, resources):
    task_resource = {}
    task_times = {}
    resource_usage = {j: [] for j in range(resources)}

    for i, (r_i, e_i, d_i) in enumerate(tasks):
        for j in range(resources):
            if model[u[i][j] - 1] > 0:
                task_resource[i] = j
        task_times[i] = [t + r_i for t in range(d_i - r_i) if model[z[i][t] - 1] > 0]
        if i in task_resource:
            resource_usage[task_resource[i]].extend(task_times[i])

    for i, (r_i, e_i, d_i) in enumerate(tasks):
        if i not in task_resource:
            print_to_console_and_log(f"Lỗi: Tác vụ {i+1} không được gán cho tài nguyên nào")
            return False
        if not task_times[i] or task_times[i][0] < r_i:
            print_to_console_and_log(f"Lỗi: Tác vụ {i+1} bắt đầu trước thời gian phát hành")
            return False
        if task_times[i][-1] >= d_i:
            print_to_console_and_log(f"Lỗi: Tác vụ {i+1} kết thúc sau deadline")
            return False
        if len(task_times[i]) != e_i or any(task_times[i][k+1] - task_times[i][k] != 1 for k in range(len(task_times[i])-1)):
            print_to_console_and_log(f"Lỗi: Tác vụ {i+1} không thực thi liên tục hoặc không đúng thời gian thực thi")
            return False

    for j, times in resource_usage.items():
        if len(times) != len(set(times)):
            print_to_console_and_log(f"Lỗi: Tài nguyên {j} được sử dụng bởi nhiều tác vụ cùng lúc")
            return False

    print_to_console_and_log("Giải pháp hợp lệ!")
    return True

def solve_with_timeout(tasks, resources, result_container, finished_event):
    global sat_solver
    sat_solver = Glucose3()
    try:
        print_to_console_and_log("Đang giải...")
        u, z, D = encode_problem_es3(tasks, resources)
        result = sat_solver.solve()
        print_to_console_and_log("Đã giải xong!")
        
        if result:
            model = sat_solver.get_model()
            if model:
                result_container.update({'status': 'SAT', 'model': model, 'u': u, 'z': z, 'D': D})
            else:
                result_container['status'] = 'TIMEOUT'
        else:
            result_container['status'] = 'UNSAT'
    except Exception as e:
        result_container.update({'status': 'ERROR', 'error': str(e)})
    finished_event.set()

def solve_es3(tasks, resources):
    result_container = {}
    finished_event = Event()
    
    start_time = time.time()
    solver_thread = Thread(target=solve_with_timeout, args=(tasks, resources, result_container, finished_event))
    solver_thread.start()
    
    finished = finished_event.wait(timeout=time_budget)
    solve_time = time.time() - start_time
    
    if not finished:
        sat_solver.interrupt()
        solver_thread.join()
        return "TIMEOUT", solve_time, 0, 0
    
    status = result_container.get('status')
    if status == 'SAT':
        model, u, z, D = result_container['model'], result_container['u'], result_container['z'], result_container['D']
        print_to_console_and_log("SAT")
        for i, (r_i, _, d_i) in enumerate(tasks):
            for j in range(resources):
                if model[u[i][j] - 1] > 0:
                    print_to_console_and_log(f"Tác vụ {i+1} được gán cho tài nguyên {j+1}")
            for t in range(d_i - r_i):
                if model[z[i][t] - 1] > 0:
                    print_to_console_and_log(f"Tác vụ {i+1} truy cập tài nguyên tại thời điểm {t + r_i}")
            for j in range(resources):
                for t in range(d_i - tasks[i][1] - r_i + 1):
                    if model[D[i][j][t] - 1] > 0:
                        print_to_console_and_log(f"Tác vụ {i+1} bắt đầu truy cập tài nguyên {j+1} tại thời điểm {t + r_i}")
        
        if not validate_solution(tasks, model, u, z, D, resources):
            sys.exit(1)
        
        num_vars, num_clauses = sat_solver.nof_vars(), sat_solver.nof_clauses()
        sat_solver.delete()
        return "SAT", solve_time, num_vars, num_clauses
    
    elif status == 'UNSAT':
        print_to_console_and_log("UNSAT")
        num_vars, num_clauses = sat_solver.nof_vars(), sat_solver.nof_clauses()
        sat_solver.delete()
        return "UNSAT", solve_time, num_vars, num_clauses
    else:
        print_to_console_and_log(f"Lỗi: {result_container.get('error', 'Lỗi không xác định')}")
        sat_solver.delete()
        return status or 'ERROR', solve_time, 0, 0

def process_input_files(input_folder, resources=200):
    global id_counter, type
    for filename in os.listdir(input_folder):
        if filename.endswith(".txt"):
            file_path = os.path.join(input_folder, filename)
            with open(file_path, 'r') as f:
                num_tasks = int(f.readline().strip())
                tasks = ast.literal_eval(f.readline().strip())
                print(f"Tác vụ: {tasks}")

            print(f"Đang xử lý {filename}...")
            res, solve_time, num_vars, num_clauses = solve_es3(tasks, resources)
            result_dict = {
                "ID": id_counter, "Problem": os.path.basename(filename), "Type": type,
                "Time": solve_time, "Result": res, "Variables": num_vars, "Clauses": num_clauses
            }
            write_to_xlsx(result_dict)
            id_counter += 1

# Thực thi chính
input_folder = "input/" + sys.argv[1]
process_input_files(input_folder)
log_file.close()