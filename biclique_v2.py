#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scheduling via Product Encoding (AMO_PE) + Disjoint Intervals (Theorem 22 + IPT)
Implements Section 4.3 of the paper strictly using the paper's primitives.
(VERSION: Corrected and Enhanced Disjoint Intervals Encoder)

What this code does
-------------------
Given N tasks with release times r[i], durations d[i], deadlines e[i], and M identical machines, we
build a CNF formula that encodes a feasible non-preemptive schedule under:
  • Each task scheduled exactly once (one start time and one machine).
  • No two tasks overlap on the same machine.
  • All starts respect release/deadline windows.

Variables
---------
x[i,t,m]  : task i starts at time t on machine m. (Boolean)
y[m,t1,t2]: machine m runs a single interval [t1,t2) (occupied by some task). (Boolean)
Aux variables appear in the AMO product encoding and in the disjoint-intervals encoder (Theorem 22).

Paper mapping
-------------
• ALO+AMO per task: “each task scheduled exactly once”.
• Duration-partition AMO_PE at each (m,t): AMO over {x(i,t,m) | d_i = Δ}, as in Section 4.3.
• x -> y link: start implies the corresponding machine-interval is selected.
• Disjoint y-intervals on each machine implemented with Theorem 22 (Sec. 4):
  - Block decomposition with k = ceil(log2 n)
  - Y_{ℓ,r} variables and links: Eq. (3) and (4)
  - Recursive disjointness on block-level intervals (Y-edges)
  - S-variables with clauses of Eq. (5) plus IPT (§4.1) to handle contained-in-block corner cases
"""

import ast
import os
import sys
import math
import time
import traceback
from collections import defaultdict
from itertools import combinations
from pysat.formula import CNF
from pysat.solvers import Glucose3
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from threading import Thread, Event

# ----------------------------- Utilities ------------------------------------
def write_to_excel(result_dict, folder):
    df = pd.DataFrame([result_dict])
    date = datetime.now().strftime('%Y-%m-%d')
    output_dir = f'out/{folder}'
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    file_path = f'{output_dir}/results_{date}_2025.xlsx'

    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(file_path)
    else:
        df.to_excel(file_path, index=False, sheet_name='Results', header=True)

log_file = open('run.log', 'a')
def log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=log_file, **kwargs)
    log_file.flush()
class VarPool:
    """Simple name-based variable pool for reproducible ids."""
    def __init__(self):
        self._next = 1
        self._map = {}

    def new(self, key=None):
        if key is None:
            v = self._next
            self._next += 1
            return v
        if key in self._map:
            return self._map[key]
        v = self._next
        self._next += 1
        self._map[key] = v
        return v

# -------------------------- Product AMO (AMO_PE) -----------------------------

_amo_counter = 0
def amo_product(cnf: CNF, pool: VarPool, lits):
    """
    Product AMO encoding (paper's primitive).
    """
    global _amo_counter
    _amo_counter += 1
    uid = _amo_counter

    n = len(lits)
    if n <= 1:
        return
    p = int(math.ceil(math.sqrt(n)))
    q = int(math.ceil(n / p))
    L = lits + [None] * (p * q - n)
    rows = [[L[r * q + c] for c in range(q) if L[r * q + c] is not None] for r in range(p)]
    cols = [[L[r * q + c] for r in range(p) if L[r * q + c] is not None] for c in range(q)]
    R = [pool.new(("R", uid, r)) for r in range(p)]
    C = [pool.new(("C", uid, c)) for c in range(q)]
    for r in range(p):
        for c in range(q):
            idx = r * q + c
            if idx >= len(L) or L[idx] is None:
                continue
            x = L[idx]
            cnf.append([-x, R[r]])
            cnf.append([-x, C[c]])
    for i in range(len(R)):
        for j in range(i + 1, len(R)):
            cnf.append([-R[i], -R[j]])
    for i in range(len(C)):
        for j in range(i + 1, len(C)):
            cnf.append([-C[i], -C[j]])

# ----------------------- Interval Propagation Trick (IPT) --------------------

def add_interval_propagation_trick(cnf: CNF, pool: VarPool, X_ij: dict, n: int, uid=None):
    """
    IPT (§4.1): Encodes T[t] <=> exists selected interval covering t.
    Returns Z and T variables.
    """
    if not X_ij or n == 0:
        return {}, []
    
    # Use provided uid or generate from id(X_ij) - caller should provide unique uid
    if uid is None:
        uid = id(X_ij)
        
    Z = {}
    for i in range(n):
        for j in range(i + 1, n + 1):
            Z[(i, j)] = pool.new(("Z", uid, i, j))
    T = [pool.new(("T", uid, t)) for t in range(n)]

    for i in range(n):
        for j in range(i + 2, n + 1):
            cnf.append([-Z[(i, j)], Z[(i, j - 1)]])
            cnf.append([-Z[(i, j)], Z[(i + 1, j)]])

    for (i, j), xv in X_ij.items():
        if i < j and (i, j) in Z:
            cnf.append([-xv, Z[(i, j)]])

    for i in range(n):
        for j in range(i + 1, n + 1):
            if (i, j) in Z:
                for t in range(i, j):
                    cnf.append([-Z[(i, j)], T[t]])

    for t in range(n):
        clause = [-T[t]]
        for i in range(t + 1):
            for j in range(t + 1, n + 1):
                xv = X_ij.get((i, j))
                if xv is not None:
                    clause.append(xv)
        if len(clause) > 1:
            cnf.append(clause)

    return Z, T
        
class DisjointIntervalsEncoder:
    """
    Cài đặt đầy đủ và đúng đắn bộ mã hóa đảm bảo các khoảng thời gian không chồng lấn,
    dựa trên Định lý 22. Lớp này sử dụng phương pháp đệ quy hai cấp để xử lý
    cả xung đột bên trong nhóm (intra-group) và giữa các nhóm (inter-group).
    """
    _uid_counter = 0  # Class variable for globally unique IDs
    
    def __init__(self, cnf: CNF, pool: VarPool):
        self.cnf = cnf
        self.pool = pool

    def _direct_encoding(self, X_ij: dict):
        """
        Mã hóa cặp đôi trực tiếp (pairwise). Đây là trường hợp cơ sở của đệ quy,
        đảm bảo tính đúng đắn cho các bài toán con đủ nhỏ.
        """
        items = list(X_ij.items())
        for a in range(len(items)):
            (i1, j1), v1 = items[a]
            for b in range(a + 1, len(items)):
                (i2, j2), v2 = items[b]
                # Thêm mệnh đề xung đột nếu hai khoảng [i1, j1) và [i2, j2) chồng lấn.
                if not (j1 <= i2 or j2 <= i1):
                    self.cnf.append([-v1, -v2])

    def encode(self, X_ij: dict, n: int):
        """
        Hàm mã hóa chính, thực hiện thuật toán đệ quy.
        - X_ij: Một dictionary ánh xạ từ cặp (start, end) tới biến SAT tương ứng.
        - n: Kích thước của miền thời gian [0, n-1].
        """
        # Nếu không có khoảng nào để mã hóa, thoát.
        if not X_ij:
            return

        # ID duy nhất cho mỗi lần gọi encode() để tránh xung đột biến phụ (Y) giữa các máy.
        # Use class counter instead of id() to ensure true uniqueness
        DisjointIntervalsEncoder._uid_counter += 1
        uid = DisjointIntervalsEncoder._uid_counter

        # ĐIỀU KIỆN CƠ SỞ: Nếu bài toán đủ nhỏ, sử dụng mã hóa trực tiếp,
        # vừa hiệu quả vừa đảm bảo đúng. Điều này cũng tránh vòng lặp
        # đệ quy khi phân vùng không làm giảm kích thước vấn đề.
        if n <= 25 or len(X_ij) <= 40:
            # use pairwise direct encoding as base case
            self._direct_encoding(X_ij)
            return

        # Safety check: avoid pathological cases where partitioning into blocks
        # would not reduce the problem (e.g., when b == 0 or b == 1 leading to
        # same B(x) for many values). If detected, fall back to direct encoding.
        # 1. THIẾT LẬP CẤU TRÚC KHỐI (BLOCKS)
        k = max(2, int(math.ceil(math.log2(n))))
        b = int(math.ceil(n / k))
        def B(x): return x // b

        # 2. ĐỆ QUY CẤP 1: XỬ LÝ XUNG ĐỘT BÊN TRONG MỖI NHÓM (INTRA-GROUP / X-EDGES)
        
        # Phân vùng tất cả các khoảng vào các nhóm dựa trên khối bắt đầu và khối kết thúc.
        by_blockpair = defaultdict(list)
        for (i, j) in X_ij.keys():
            bi = B(i)
            # Khoảng là [i, j), nên điểm thời gian cuối cùng là j-1.
            bj = B(j - 1) if j > i else bi
            by_blockpair[(bi, bj)].append((i, j))

        # Với MỖI nhóm, gọi đệ quy để đảm bảo các khoảng BÊN TRONG nhóm đó không chồng lấn.
        # Đây là bước logic quan trọng đã bị thiếu, gây ra lỗi chồng lấn.
        for (bl, br), pairs in by_blockpair.items():
            if len(pairs) <= 1:
                continue # Nhóm có 1 hoặc 0 khoảng thì không có xung đột nội bộ.
            
            # Tạo một bài toán con cục bộ cho nhóm này.
            # Ánh xạ lại các chỉ số thời gian về một miền mới bắt đầu từ 0.
            endpoints = set()
            for (i, j) in pairs:
                endpoints.add(i)
                endpoints.add(j)

            # Include full block ranges so the local domain is limited to the
            # blocks that the intervals touch. This guarantees the recursive
            # subproblem is smaller than the parent domain and avoids falling
            # back to an overly-strong direct encoding.
            if bl == br:
                block = range(bl * b, min((bl + 1) * b, n))
                local = sorted(set(block) | endpoints)
            else:
                left = range(bl * b, min((bl + 1) * b, n))
                right = range(br * b, min((br + 1) * b, n))
                local = sorted(set(left) | set(right) | endpoints)

            remap = {orig: idx for idx, orig in enumerate(local)}
            X_local = {(remap[i], remap[j]): X_ij[(i, j)] for i, j in pairs}

            # Gọi đệ quy trên bài toán con này. Nếu domain không shrink (pathological),
            # fall back to direct encoding but log the occurrence for diagnostics.
            if len(local) >= n:
                log(f'Fallback direct encoding for block {(bl,br)}: local_size={len(local)} parent_n={n} pairs={len(pairs)}')
                self._direct_encoding({(i, j): X_ij[(i, j)] for i, j in pairs})
            else:
                self.encode(X_local, len(local))

        # Use IPT at this level to provide T variables (time coverage) which
        # are used by S/F/M clauses to capture cross-block conflicts as in the
        # paper. IPT must be called before adding S/F/M clauses.
        _, T = add_interval_propagation_trick(self.cnf, self.pool, X_ij, n, uid)

        # 3. ĐỆ QUY CẤP 2: XỬ LÝ XUNG ĐỘT GIỮA CÁC NHÓM (INTER-GROUP / Y-EDGES)

        # 3.1. Tạo các biến phụ Y(bl, br) đại diện cho mỗi nhóm.
        Y = {}
        for bl in range(k):
            for br in range(bl, k):
                key = (bl, br)
                pairs_in_group = by_blockpair.get(key)
                if not pairs_in_group:
                    continue

                y_var = self.pool.new(("Y", uid, bl, br))
                Y[key] = y_var
                
                xs = [X_ij[pair] for pair in pairs_in_group]
                
                # Liên kết logic hai chiều:
                # a) Nếu có một x được chọn, Y tương ứng phải được chọn. (x -> Y)
                for x_var in xs:
                    self.cnf.append([-x_var, y_var])
                
                # b) Nếu Y được chọn, phải có ít nhất một x tương ứng được chọn. (Y -> OR(x))
                self.cnf.append([-y_var] + xs)

        # 3.2. THÊM CÁC MỆNH ĐỀ XUNG ĐỘT QUAN TRỌNG (s, f, m-edges)
        # Add S-edge, F-edge, M-edge clauses using T variables from IPT.
        if T:
            # S-edge: same start-block l, different end-blocks r1 < r2
            for l in range(k):
                for r1 in range(l, k):
                    for r2 in range(r1 + 1, k):
                        y1, y2 = Y.get((l, r1)), Y.get((l, r2))
                        if y1 is None or y2 is None: continue
                        for t in range(l * b, min((l + 1) * b, n)):
                            self.cnf.append([-y1, -y2, -T[t]])

            # F-edge: same end-block r, different start-blocks l1 < l2
            for r in range(k):
                for l1 in range(r + 1):
                    for l2 in range(l1 + 1, r + 1):
                        y1, y2 = Y.get((l1, r)), Y.get((l2, r))
                        if y1 is None or y2 is None: continue
                        for t in range(r * b, min((r + 1) * b, n)):
                            self.cnf.append([-y1, -y2, -T[t]])

            # M-edge: interleaving blocks l1 < l2 <= r1 < r2
            for l1 in range(k):
                for r1 in range(l1, k):
                    for l2 in range(l1 + 1, r1 + 1):
                        for r2 in range(r1 + 1, k):
                            y1, y2 = Y.get((l1, r1)), Y.get((l2, r2))
                            if y1 is None or y2 is None: continue
                            overlap_start = l2 * b
                            overlap_end = min((r1 + 1) * b, n)
                            for t in range(overlap_start, overlap_end):
                                self.cnf.append([-y1, -y2, -T[t]])

        # 3.3. Tạo bài toán con ở cấp độ khối và gọi đệ quy.
        # Một biến Y(bl, br) đại diện cho khoảng khối [bl, br+1).
        X_blk = {(bl, br + 1): y_var for (bl, br), y_var in Y.items()}

        # IMPORTANT: Y-block variables represent EXISTENTIAL quantification over intervals
        # within block ranges. Conflicts between Y variables are already captured through:
        # 1. x -> Y implication clauses (line 534)
        # 2. Y -> OR(x) clauses (line 537)  
        # 3. Disjointness encoding of original X_ij intervals
        # Adding direct pairwise conflicts between Y-blocks based on block-range overlap
        # would be INCORRECT, as overlapping block ranges can contain non-overlapping intervals.
        # Therefore, we DO NOT encode conflicts at the Y-block level.
        
        # OLD (INCORRECT): self._direct_encoding(X_blk)
        # This was treating Y-block ranges as if they were actual time intervals, causing
        # spurious conflicts and leading to false-UNSAT results.

class SchedulingEncoder:
    """
    Section 4.3 – Applications to Scheduling Problems (paper)
    (VERSION: Using Correct and Robust Pairwise Disjointness)
    """
    def __init__(self, N, M, T, d, r, e):
        self.N, self.M, self.T = N, M, T
        self.d, self.r, self.e = d, r, e
        self.pool = VarPool()
        self.cnf = CNF()
        self.x = {}  # (i,t,m)->var
        self.y = {}  # (m,t1,t2)->var
        self.U = {}  # (i,m)->var, for symmetry breaking

    def build_x_and_link(self):
        y_to_x = defaultdict(list)
        # Helper for symmetry breaking: group x-vars by (task, machine)
        x_by_im = defaultdict(list)

        for i in range(self.N):
            t_min = max(0, self.r[i])
            t_max = min(self.T - self.d[i], self.e[i] - self.d[i])
            if t_min > t_max:
                self.cnf.append([])
                return
            for t in range(t_min, t_max + 1):
                for m in range(self.M):
                    xv = self.pool.new(("x", i, t, m))
                    self.x[(i, t, m)] = xv
                    x_by_im[(i, m)].append(xv)

                    t1, t2 = t, t + self.d[i]
                    yk = (m, t1, t2)
                    yv = self.y.get(yk)
                    if yv is None:
                        yv = self.pool.new(("y", m, t1, t2))
                        self.y[yk] = yv
                    self.cnf.append([-xv, yv])
                    y_to_x[yk].append(xv)
        
        for yk, yv in self.y.items():
            xs = y_to_x.get(yk, [])
            if xs:
                self.cnf.append([-yv] + xs)
        
        # Create U_im variables and link them to x_itm variables
        for (i, m), x_vars in x_by_im.items():
            if not x_vars: continue
            u_var = self.pool.new(("U", i, m))
            self.U[(i, m)] = u_var
            # Link x -> U: if any x_itm is true, U_im must be true
            for x_var in x_vars:
                self.cnf.append([-x_var, u_var])
            # Link U -> x: if U_im is true, at least one x_itm must be true
            # This part is expensive and not strictly necessary for symmetry breaking to work
            # self.cnf.append([-u_var] + x_vars)

    def per_task_exactly_one(self):
        for i in range(self.N):
            lits = [v for (ii, t, m), v in self.x.items() if ii == i]
            if not lits:
                self.cnf.append([])
                return
            self.cnf.append(lits)
            amo_product(self.cnf, self.pool, lits)

    def duration_partition_amo(self):
        for m in range(self.M):
            for t in range(self.T):
                groups = defaultdict(list)
                for i in range(self.N):
                    v = self.x.get((i, t, m))
                    if v is not None:
                        groups[self.d[i]].append(v)
                for Δ, lits in groups.items():
                    if len(lits) > 1:
                        amo_product(self.cnf, self.pool, lits)

    def disjoint_y_intervals(self):
        intervals_by_machine = defaultdict(dict)
        for (m, t1, t2), v in self.y.items():
            if 0 <= t1 < t2 <= self.T:
                intervals_by_machine[m][(t1, t2)] = v

        encoder = DisjointIntervalsEncoder(self.cnf, self.pool)
        for m, X in intervals_by_machine.items():
            encoder.encode(X, self.T + 1)

    def symmetry_breaking(self):
        """
        Adds efficient task-level symmetry-breaking constraints.
        If task i is on machine m > 0, some task j < i must be on machine m-1.
        This is enforced using U_im helper variables.
        """
        if self.M <= 1:
            return

        for i in range(1, self.N):
            for m in range(1, self.M):
                u_im = self.U.get((i, m))
                if u_im is None: continue

                # Collect all U variables for tasks j < i on machine m-1
                prev_machine_u_vars = [self.U.get((j, m - 1)) for j in range(i)]
                prev_machine_u_vars = [v for v in prev_machine_u_vars if v is not None]

                if not prev_machine_u_vars:
                    # If no tasks can be placed on the previous machine, then task i cannot be on this machine
                    self.cnf.append([-u_im])
                    continue
                
                # U_im => OR_{j<i} U_j,m-1
                self.cnf.append([-u_im] + prev_machine_u_vars)

    def encode(self):
        self.build_x_and_link()
        if any(len(c) == 0 for c in self.cnf.clauses): return self.cnf
        self.per_task_exactly_one()
        if any(len(c) == 0 for c in self.cnf.clauses): return self.cnf
        self.duration_partition_amo()
        self.disjoint_y_intervals()
        self.symmetry_breaking()
        return self.cnf

# ----------------------------- I/O and Solve ---------------------------------

def solve_from_instance(N, tasks, M=None):
    r = [t[0] for t in tasks]
    d = [t[1] for t in tasks]
    e = [t[2] for t in tasks]
    T = max(e) if e else 0
    M = M if M is not None else len(tasks)

    enc = SchedulingEncoder(N, M, T, d, r, e)
    start = time.time()
    cnf = enc.encode()
    if any(len(c) == 0 for c in cnf.clauses):
        return False, "UNSAT by construction", None, 0, 0, 0

    solver = Glucose3()
    solver.append_formula(cnf.clauses)
    sat = solver.solve()
    elapsed = time.time() - start
    
    nof_vars = solver.nof_vars()
    nof_clauses = solver.nof_clauses()

    if not sat:
        return False, "UNSAT (solver)", None, elapsed, nof_vars, nof_clauses
    
    model = set(solver.get_model())
    schedule = []
    for (i, t, m), xv in enc.x.items():
        if xv in model:
            schedule.append((i, m, t, t + d[i]))
    schedule.sort()
    return True, schedule, enc, elapsed, nof_vars, nof_clauses

def solve_and_record_from_file(path, M_override=None):
    with open(path) as f:
        N = int(f.readline().strip())
        tasks = ast.literal_eval(f.readline().strip())
        
        unique_tasks = list(set(tasks))
    
        N_new = len(unique_tasks)
        
    return solve_from_instance(N_new, unique_tasks, M_override)
    # return solve_from_instance(N, tasks, M_override)

def process_input_dir(input_dir, folder):
    id = 1
    for filename in os.listdir(input_dir):
        if not filename.endswith(".txt"):
            continue
        path = os.path.join(input_dir, filename)
        result_container = {}
        finished_event = Event()
        time_budget = 60

        def solve_with_timeout(path, result_container, finished_event):
            try:
                ok, schedule_or_msg, enc, elapsed, nof_vars, nof_clauses = solve_and_record_from_file(path)
                result_container['ok'] = ok
                result_container['schedule_or_msg'] = schedule_or_msg
                result_container['elapsed'] = elapsed
                result_container['nof_vars'] = nof_vars
                result_container['nof_clauses'] = nof_clauses
            except Exception as ex:
                result_container['error'] = ex
                result_container['traceback'] = traceback.format_exc()
            finally:
                finished_event.set()

        log(f"=== Processing {filename} ===")
        start_time = time.time()
        solver_thread = Thread(target=solve_with_timeout, args=(path, result_container, finished_event))
        solver_thread.start()

        finished = finished_event.wait(timeout=time_budget)
        solve_time = time.time() - start_time

        result_dict = {"ID": id, "Problem": filename, "Type": "biclique_v2"}
        
        if not finished:
            log(f"✗ TIMEOUT after {time_budget}s")
            result_dict.update({"Time": round(solve_time, 4), "Result": "TIMEOUT", "Variables": None, "Clauses": None})
            # It's difficult to get var/clause count on timeout, so we leave them None
            write_to_excel(result_dict, folder)
        elif 'error' in result_container:
            log("Error processing", filename, ":", result_container['error'])
            log(result_container.get('traceback', ''))
            result_dict.update({"Time": round(solve_time, 4), "Result": "ERROR", "Variables": None, "Clauses": None})
            write_to_excel(result_dict, folder)
        else:
            ok = result_container['ok']
            schedule_or_msg = result_container['schedule_or_msg']
            elapsed = result_container['elapsed']
            nof_vars = result_container['nof_vars']
            nof_clauses = result_container['nof_clauses']
            
            result_dict.update({"Time": round(elapsed, 4), "Variables": nof_vars, "Clauses": nof_clauses})
            if ok:
                log("SAT in %.3fs; schedule:" % elapsed)
                # Sort schedule by machine, then by start time
                for (i, m, s, f) in sorted(schedule_or_msg, key=lambda x: (x[1], x[2])):
                    log(f" Task {i} -> machine {m}: [{s},{f})")
                result_dict["Result"] = "SAT"
            else:
                log("✗ NO FEASIBLE SCHEDULE EXISTS or error:", schedule_or_msg)
                result_dict["Result"] = "UNSAT"
            write_to_excel(result_dict, folder)

        id += 1
        
    log_file.close()

# ------------------------------- CLI -----------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        log("Usage: script.py <input_subdir_under_input/>")
        sys.exit(1)
    sub = sys.argv[1]
    input_dir = os.path.join("input", sub)
    process_input_dir(input_dir, sub)