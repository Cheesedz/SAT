#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scheduling via Product Encoding (AMO_PE) + Disjoint Intervals (Theorem 22 + IPT)
Implements Section 4.3 of the paper strictly using the paper's primitives.
(VERSION: Corrected and Enhanced Disjoint Intervals Encoder)

What this code does
-------------------
Given N tasks with release times r[i], durations d[i], deadlines e[i], and M identical machines, we
build a CNF formula that encodes a feasible non-preemptive schedule under:
  • Each task scheduled exactly once (one start time and one machine).
  • No two tasks overlap on the same machine.
  • All starts respect release/deadline windows.

Variables
---------
x[i,t,m]  : task i starts at time t on machine m. (Boolean)
y[m,t1,t2]: machine m runs a single interval [t1,t2) (occupied by some task). (Boolean)
Aux variables appear in the AMO product encoding and in the disjoint-intervals encoder (Theorem 22).

Paper mapping
-------------
• ALO+AMO per task: “each task scheduled exactly once”.
• Duration-partition AMO_PE at each (m,t): AMO over {x(i,t,m) | d_i = Δ}, as in Section 4.3.
• x -> y link: start implies the corresponding machine-interval is selected.
• Disjoint y-intervals on each machine implemented with Theorem 22 (Sec. 4):
  - Block decomposition with k = ceil(log2 n)
  - Y_{ℓ,r} variables and links: Eq. (3) and (4)
  - Recursive disjointness on block-level intervals (Y-edges)
  - S-variables with clauses of Eq. (5) plus IPT (§4.1) to handle contained-in-block corner cases
"""

import ast
import os
import sys
import math
import time
import traceback
from collections import defaultdict
from itertools import combinations
from pysat.formula import CNF
from pysat.solvers import Cadical195
from pysat.solvers import Glucose3
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from threading import Thread, Event

try:
    from pypblib import pblib
    from pypblib.pblib import PBConfig, Pb2cnf
    PYPBLIB_AVAILABLE = True
except ImportError:
    print("Warning: pypblib not available. Install with: pip install pypblib")
    print("Falling back to standard AMO encoding (slower).")
    PYPBLIB_AVAILABLE = False

# ----------------------------- Utilities ------------------------------------
def write_to_excel(result_dict, folder):
    df = pd.DataFrame([result_dict])
    date = datetime.now().strftime('%Y-%m-%d')
    output_dir = f'out/{folder}'
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    file_path = f'{output_dir}/biclique_pb/results_{date}_2025.xlsx'

    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(file_path)
    else:
        df.to_excel(file_path, index=False, sheet_name='Results', header=True)

log_file = open('run.log', 'a')
def log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=log_file, **kwargs)
    log_file.flush()
class VarPool:
    """Simple name-based variable pool for reproducible ids."""
    def __init__(self):
        self._next = 1
        self._map = {}

    def new(self, key=None):
        if key is None:
            v = self._next
            self._next += 1
            return v
        if key in self._map:
            return self._map[key]
        v = self._next
        self._next += 1
        self._map[key] = v
        return v

# -------------------------- Product AMO (AMO_PE) -----------------------------

_amo_counter = 0
def sequential_amo(cnf: CNF, pool: VarPool, lits):
    """
    Sequential (counter) AMO encoding.
    Uses O(n) auxiliary variables and O(n) clauses.
    
    For literals [x1, x2, ..., xn], creates auxiliary variables [s1, s2, ..., s_{n-1}]
    with clauses:
    - ¬x1 ∨ s1       (if x1 is true, counter starts)
    - ¬xi ∨ si       (if xi is true, counter is at least i)
    - ¬si-1 ∨ si     (counter is monotonic)
    - ¬xi ∨ ¬si-1    (if counter already at i-1, xi cannot be true)
    
    This ensures at most one literal can be true.
    """
    n = len(lits)
    if n <= 1:
        return
    
    # Create auxiliary variables s1, s2, ..., s_{n-1}
    s = [pool.new(("seq_amo", id(lits), i)) for i in range(n - 1)]
    
    # First literal: ¬x1 ∨ s1
    cnf.append([-lits[0], s[0]])
    
    # Middle literals
    for i in range(1, n - 1):
        cnf.append([-lits[i], s[i]])           # ¬xi ∨ si
        cnf.append([-s[i-1], s[i]])            # ¬si-1 ∨ si (monotonicity)
        cnf.append([-lits[i], -s[i-1]])        # ¬xi ∨ ¬si-1 (conflict)
    
    # Last literal: ¬xn ∨ ¬s_{n-1}
    cnf.append([-lits[n-1], -s[n-2]])
    
def pairwise_amo(cnf: CNF, lits):
    """Simple pairwise AMO encoding."""
    for i in range(len(lits)):
        for j in range(i + 1, len(lits)):
            cnf.append([-lits[i], -lits[j]])

# ----------------------- Interval Propagation Trick (IPT) --------------------

def add_interval_propagation_trick(cnf: CNF, pool: VarPool, X_ij: dict, n: int, uid=None):
    """
    IPT (§4.1): Encodes T[t] <=> exists selected interval covering t.
    Returns Z and T variables.
    """
    if not X_ij or n == 0:
        return {}, []
    
    # Use provided uid or generate from id(X_ij) - caller should provide unique uid
    if uid is None:
        uid = id(X_ij)
        
    Z = {}
    for i in range(n):
        for j in range(i + 1, n + 1):
            Z[(i, j)] = pool.new(("Z", uid, i, j))
    T = [pool.new(("T", uid, t)) for t in range(n)]

    for i in range(n):
        for j in range(i + 2, n + 1):
            cnf.append([-Z[(i, j)], Z[(i, j - 1)]])
            cnf.append([-Z[(i, j)], Z[(i + 1, j)]])

    for (i, j), xv in X_ij.items():
        if i < j and (i, j) in Z:
            cnf.append([-xv, Z[(i, j)]])

    for i in range(n):
        for j in range(i + 1, n + 1):
            if (i, j) in Z:
                for t in range(i, j):
                    cnf.append([-Z[(i, j)], T[t]])

    for t in range(n):
        clause = [-T[t]]
        for i in range(t + 1):
            for j in range(t + 1, n + 1):
                xv = X_ij.get((i, j))
                if xv is not None:
                    clause.append(xv)
        if len(clause) > 1:
            cnf.append(clause)

    return Z, T
        
class DisjointIntervalsEncoder:
    _uid_counter = 0

    def __init__(self, cnf: CNF, pool: VarPool):
        self.cnf = cnf
        self.pool = pool

    def encode(self, X_ij: dict, n: int):
        """
        Mã hóa tính chất các khoảng rời nhau (Disjoint)[cite: 300].
        Độ phức tạp: O(n^2 log n).
        """
        if not X_ij: return
        DisjointIntervalsEncoder._uid_counter += 1
        uid = DisjointIntervalsEncoder._uid_counter

        # --- MEMORY OPTIMIZATION ---
        # Nếu số lượng khoảng ít hoặc không gian thời gian nhỏ, dùng Pairwise O(N^2) clauses
        # nhưng 0 biến phụ. Nhanh hơn và nhẹ RAM hơn nhiều so với đệ quy block.
        # Ngưỡng 64 là con số thực nghiệm tốt cho bài toán này.
        if len(X_ij) < 64 or n < 64: 
            items = list(X_ij.items())
            for a in range(len(items)):
                (i1, j1), v1 = items[a]
                for b in range(a + 1, len(items)):
                    (i2, j2), v2 = items[b]
                    # Kiểm tra chồng lấn: max_start < min_end
                    if max(i1, i2) < min(j1, j2): 
                        self.cnf.append([-v1, -v2])
            return

        # 1. Block Decomposition [cite: 277]
        k = max(2, int(math.ceil(math.log2(n)))) # Số lượng block
        b = int(math.ceil(n / k))              # Kích thước block
        def B(x): return min(x // b, k - 1)

        # Phân loại cạnh (Edges)
        by_blockpair = defaultdict(list)
        for (i, j) in X_ij.keys():
            # Interval [i, j), thời điểm cuối là j-1
            bi = B(i)
            bj = B(j - 1) if j > i else bi
            by_blockpair[(bi, bj)].append((i, j))

        # 2. Intra-block constraints (Đệ quy)
        # Tương ứng với x-edges trong bài báo [cite: 281, 308]
        for (bl, br), pairs in by_blockpair.items():
            if len(pairs) <= 1: continue
            
            # Tạo không gian con
            sub_X = {p: X_ij[p] for p in pairs}
            
            # Nếu block pair này chỉ nằm gọn trong 1 hoặc 2 block thực tế,
            # gọi đệ quy với không gian thời gian thu hẹp
            start_t = bl * b
            end_t = min((br + 1) * b, n)
            local_len = end_t - start_t
            
            # Mapping lại index về 0..local_len
            mapped_X = {}
            for (i, j), val in sub_X.items():
                mapped_X[(i - start_t, j - start_t)] = val

            # Safety: if the local domain did not shrink (pathological case),
            # fall back to direct pairwise encoding to avoid infinite recursion.
            # Compare against sub_X (the group) rather than the whole problem X_ij.
            if local_len >= n or len(mapped_X) >= len(sub_X):
                items = list(sub_X.items())
                for a in range(len(items)):
                    (i1, j1), v1 = items[a]
                    for b in range(a + 1, len(items)):
                        (i2, j2), v2 = items[b]
                        if max(i1, i2) < min(j1, j2):
                            self.cnf.append([-v1, -v2])
            else:
                self.encode(mapped_X, local_len)

        # 3. Inter-block constraints (y-edges, s-edges, f-edges, m-edges)
        # Sử dụng biến T (Time occupancy) để xử lý các va chạm giữa các nhóm
        Z, T = add_interval_propagation_trick(self.cnf, self.pool, X_ij, n, uid)
        
        # Tạo biến Y_{l,r}: Có ít nhất 1 interval thuộc nhóm block (l, r) được chọn [cite: 311]
        Y = {}
        for key, pairs in by_blockpair.items():
            y_var = self.pool.new(("Y", uid, key))
            Y[key] = y_var
            # Logic: Y_{l,r} <-> OR(x_{i,j} in group)
            # Clause: -x -> Y
            for p in pairs:
                self.cnf.append([-X_ij[p], y_var])
            # Clause: -Y -> OR(x) (Optional for correctness but helps propagation)
            # self.cnf.append([-y_var] + [X_ij[p] for p in pairs]) 

        # Xử lý xung đột giữa các Y dựa trên va chạm khối
        y_items = list(Y.items())
        for idx1 in range(len(y_items)):
            (l1, r1), y1 = y_items[idx1]
            for idx2 in range(idx1 + 1, len(y_items)):
                (l2, r2), y2 = y_items[idx2]
                
                # Xác định khoảng thời gian giao nhau của 2 nhóm BLOCK
                # Nhóm 1 bao phủ từ block l1 đến r1 (thời gian: l1*b -> (r1+1)*b)
                start1, end1 = l1 * b, (r1 + 1) * b
                start2, end2 = l2 * b, (r2 + 1) * b
                
                overlap_start = max(start1, start2)
                overlap_end = min(end1, end2)
                
                if overlap_start < overlap_end:
                    # Có xung đột về mặt không gian thời gian.
                    # Dùng biến T để bắt buộc: Nếu Y1 và Y2 cùng bật, 
                    # thì không được có bất kỳ thời điểm t nào trong vùng giao bị trùng.
                    
                    # Optimization: Nếu 2 nhóm block hoàn toàn bao phủ nhau (ví dụ nằm lồng nhau chặt),
                    # ta có thể xung đột trực tiếp Y1 + Y2.
                    # Nhưng để an toàn và tổng quát (cover S, F, M edges), ta dùng T.
                    
                    # Chỉ sinh clause cho vùng giao
                    for t in range(overlap_start, min(overlap_end, n)):
                        # Guard: T should have entries for 0..n-1. If for some reason
                        # T is shorter (pathological), skip that timepoint instead
                        # of raising IndexError.
                        if t < len(T):
                            self.cnf.append([-y1, -y2, -T[t]])

class SchedulingEncoder:
    """
    Section 4.3 – Applications to Scheduling Problems (paper)
    (VERSION: Using Correct and Robust Pairwise Disjointness)
    """
    def __init__(self, N, M, T, d, r, e):
        self.N, self.M, self.T = N, M, T
        self.d, self.r, self.e = d, r, e
        self.pool = VarPool()
        self.cnf = CNF()
        self.x = {}  # (i,t,m)->var
        self.y = {}  # (m,t1,t2)->var
        self.x_by_task = defaultdict(list)  # For per_task_exactly_one
        self.x_by_machine_and_time = defaultdict(list)
        
        if PYPBLIB_AVAILABLE:
            self.pb_config = PBConfig()
            self.pb_config.set_PB_Encoder(pblib.PB_BDD)
        else:
            self.pb_config = None

    def build_x_and_link(self):

        cnf_counter = 0
        for i in range(self.N):
            t_min = max(0, self.r[i])
            t_max = min(self.T - self.d[i], self.e[i] - self.d[i])
            if t_min > t_max:
                self.cnf.append([])
                cnf_counter += 1
                return
        
            allowed_machines = range(min(i + 1, self.M))
        
            for m in allowed_machines:
                for t in range(t_min, t_max + 1):
                    xv = self.pool.new(("x", i, t, m))
                    self.x[(i, t, m)] = xv
                    self.x_by_task[i].append(xv)
                    self.x_by_machine_and_time[(m, t)].append((i, xv))  # Store task 'i' as well for duration_partition_amo

                    t1, t2 = t, t + self.d[i]
                    yk = (m, t1, t2)
                    yv = self.y.get(yk)
                    if yv is None:
                        yv = self.pool.new(("y", m, t1, t2))
                        self.y[yk] = yv
                    self.cnf.append([-xv, yv])
                    cnf_counter += 1
    
    def per_task_exactly_one(self):
        for i in range(self.N):
            lits = self.x_by_task.get(i, [])
            if not lits:
                self.cnf.append([])
                return
            self.cnf.append(lits)
            sequential_amo(self.cnf, self.pool, lits)
            # amo_product(self.cnf, self.pool, lits)

    def duration_partition_amo(self):
        # Iterate through the pre-grouped structure
        for (m, t), tasks_at_time in self.x_by_machine_and_time.items():
            groups = defaultdict(list)
            # Group by duration
            for i, v in tasks_at_time:
                groups[self.d[i]].append(v)
            
            for Δ, lits in groups.items():
                if len(lits) > 1:
                    # Use PB_BDD for at-most-one (sum <= 1)
                    if PYPBLIB_AVAILABLE and self.pb_config:
                        pb2cnf = Pb2cnf(self.pb_config)
                        formula = []
                        max_var = pb2cnf.encode_at_most_k(lits, 1, formula, self.pool._next)
                        self.pool._next = max_var + 1
                        for clause in formula:
                            self.cnf.append(clause)
                    else:
                        # Fallback to pairwise
                        pairwise_amo(self.cnf, lits)

    def disjoint_y_intervals(self):
        intervals_by_machine = defaultdict(dict)
        for (m, t1, t2), v in self.y.items():
            if 0 <= t1 < t2 <= self.T:
                intervals_by_machine[m][(t1, t2)] = v

        encoder = DisjointIntervalsEncoder(self.cnf, self.pool)
        for m, X in intervals_by_machine.items():
            encoder.encode(X, self.T + 1)
        

    def encode(self):
        self.build_x_and_link()
        print(f"Current CNF size after x,y linking: {len(self.cnf.clauses)} clauses, variables: {self.pool._next - 1}")
        if any(len(c) == 0 for c in self.cnf.clauses): return self.cnf
        self.per_task_exactly_one()
        print(f"Current CNF size after per-task exactly-one: {len(self.cnf.clauses)} clauses, variables: {self.pool._next - 1}")
        if any(len(c) == 0 for c in self.cnf.clauses): return self.cnf
        self.duration_partition_amo()
        print(f"Current CNF size after duration-partition AMO: {len(self.cnf.clauses)} clauses, variables: {self.pool._next - 1}")
        self.disjoint_y_intervals()
        print(f"Current CNF size after disjoint y-intervals: {len(self.cnf.clauses)} clauses, variables: {self.pool._next - 1}")
        return self.cnf

# ----------------------------- I/O and Solve ---------------------------------

def solve_from_instance(N, tasks, M=None):
    r = [t[0] for t in tasks]
    d = [t[1] for t in tasks]
    e = [t[2] for t in tasks]
    T = max(e) if e else 0
    M = M if M is not None else len(tasks)

    enc = SchedulingEncoder(N, M, T, d, r, e)
    start = time.time()
    cnf = enc.encode()
    if any(len(c) == 0 for c in cnf.clauses):
        return False, "UNSAT by construction", None, 0, 0, 0

    solver = Cadical195()
    solver.append_formula(cnf.clauses)
    sat = solver.solve()
    elapsed = time.time() - start
    
    nof_vars = solver.nof_vars()
    nof_clauses = solver.nof_clauses()

    if not sat:
        return False, "UNSAT (solver)", None, elapsed, nof_vars, nof_clauses
    
    model_list = solver.get_model() or []
    model = set(model_list)
    schedule = []
    for (i, t, m), xv in enc.x.items():
        if xv in model:
            schedule.append((i, m, t, t + d[i]))
    schedule.sort()
    return True, schedule, enc, elapsed, nof_vars, nof_clauses

def solve_and_record_from_file(path, M_override=None):
    with open(path) as f:
        N = int(f.readline().strip())
        tasks = ast.literal_eval(f.readline().strip())
        
    return solve_from_instance(N, tasks, M_override)

def process_input_dir(input_dir, folder):
    id = 1
    for filename in os.listdir(input_dir):
        if not filename.endswith(".txt"):
            continue
        path = os.path.join(input_dir, filename)
        result_container = {}
        finished_event = Event()
        time_budget = 600

        def solve_with_timeout(path, result_container, finished_event):
            try:
                ok, schedule_or_msg, enc, elapsed, nof_vars, nof_clauses = solve_and_record_from_file(path)
                result_container['ok'] = ok
                result_container['schedule_or_msg'] = schedule_or_msg
                result_container['elapsed'] = elapsed
                result_container['nof_vars'] = nof_vars
                result_container['nof_clauses'] = nof_clauses
            except Exception as ex:
                result_container['error'] = ex
                result_container['traceback'] = traceback.format_exc()
            finally:
                finished_event.set()

        log(f"=== Processing {filename} ===")
        start_time = time.time()
        solver_thread = Thread(target=solve_with_timeout, args=(path, result_container, finished_event))
        solver_thread.start()

        finished = finished_event.wait(timeout=time_budget)
        solve_time = time.time() - start_time

        result_dict = {"ID": id, "Problem": filename, "Type": "biclique_improved_amo_sequential"}
        
        if not finished:
            log(f"✗ TIMEOUT after {time_budget}s")
            result_dict.update({"Time": round(solve_time, 4), "Result": "TIMEOUT", "Variables": None, "Clauses": None})
            write_to_excel(result_dict, folder)
        elif 'error' in result_container:
            log("Error processing", filename, ":", result_container['error'])
            log(result_container.get('traceback', ''))
            result_dict.update({"Time": round(solve_time, 4), "Result": "ERROR", "Variables": None, "Clauses": None})
            write_to_excel(result_dict, folder)
        else:
            ok = result_container['ok']
            schedule_or_msg = result_container['schedule_or_msg']
            elapsed = result_container['elapsed']
            nof_vars = result_container['nof_vars']
            nof_clauses = result_container['nof_clauses']
            
            result_dict.update({"Time": round(elapsed, 4), "Variables": nof_vars, "Clauses": nof_clauses})
            if ok:
                log("SAT in %.3fs; schedule:" % elapsed)
                # Sort schedule by machine, then by start time
                for (i, m, s, f) in sorted(schedule_or_msg, key=lambda x: (x[1], x[2])):
                    log(f" Task {i} -> machine {m}: [{s},{f})")
                result_dict["Result"] = "SAT"
            else:
                log("✗ NO FEASIBLE SCHEDULE EXISTS or error:", schedule_or_msg)
                result_dict["Result"] = "UNSAT"
            write_to_excel(result_dict, folder)

        id += 1
        
    log_file.close()

# ------------------------------- CLI -----------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        log("Usage: script.py <input_subdir_under_input/>")
        sys.exit(1)
    sub = sys.argv[1]
    input_dir = os.path.join("input", sub)
    process_input_dir(input_dir, sub)