import ast, os, time, sys
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from pysat.solvers import Cadical153 as Cadical
from pysat.formula import CNF

log_file = open('run.log', 'a')
def log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=log_file, **kwargs)
    log_file.flush()

def write_to_excel(result_dict):
    df = pd.DataFrame([result_dict])
    date = datetime.now().strftime('%Y-%m-%d')
    output_dir = 'out'
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    file_path = f'{output_dir}/results_{date}_biclique.xlsx'
    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(file_path)
    else:
        df.to_excel(file_path, index=False, sheet_name='Results', header=True)

# ----- variable indexing (safe, collision-free) -----
def var_x(i, t, m, N, M, T):
    # times 0..T (use side = T+1)
    # index block per job: (T+1)*M
    return 1 + i * (M * (T + 1)) + t * M + m

def y_offset(N, M, T):
    return 1 + N * (M * (T + 1))

def var_y(m, t1, t2, N, M, T):
    # pack (t1,t2) with side = T+1
    side = T + 1
    return y_offset(N, M, T) + m * (side * side) + (t1 * side + t2)

# ----- encoding -----
def encode_schedule_problem(N, M, T, d, r, e):
    cnf = CNF()
    active_y_nodes = set()  # (m, t1, t2, yv)

    if T < 0:
        return cnf

    # sanity: any job with no feasible start => UNSAT instance
    for i in range(N):
        dur = d[i]
        t_start = max(0, r[i])
        t_end_allowed = min(e[i] - dur, T - dur)
        if t_start > t_end_allowed:
            # no slot for job i -> instance infeasible
            cnf.append([])  # single empty clause = UNSAT
            return cnf

    # x -> y and collect feasible x's O(M*N*T)
    bucket = defaultdict(list)
    x_lits_per_job = [[] for _ in range(N)]
    for i in range(N):
        dur = d[i]
        t_start = max(0, r[i])
        t_end_allowed = min(e[i] - dur, T - dur)
        for t1 in range(t_start, t_end_allowed + 1):
            t2 = t1 + dur
            for m in range(M):
                xi = var_x(i, t1, m, N, M, T)
                yi = var_y(m, t1, t2, N, M, T)
                # x(i,t,m) -> y(m,t,t+dur)
                cnf.append([-xi, yi])
                x_lits_per_job[i].append(xi)
                bucket[(m, t1, t2)].append(xi) 
                active_y_nodes.add((m, t1, t2, yi))
                
    def add_at_most_one(cnf, lits):
        n = len(lits)
        if n <= 1:
            return
        # sequential counter vars
        s = [cnf.nv + i + 1 for i in range(n-1)]
        cnf.nv += n-1

        # x1 -> s1
        cnf.append([-lits[0], s[0]])
        for i in range(1, n-1):
            # xi+1 -> si
            cnf.append([-lits[i], s[i]])
            # si-1 -> si
            cnf.append([-s[i-1], s[i]])
            # xi+1 & si-1 -> false
            cnf.append([-lits[i], -s[i-1]])
        # xn & sn-1 -> false
        cnf.append([-lits[n-1], -s[n-2]])

    
    # per-job exactly-one
    for i, lits in enumerate(x_lits_per_job):
        cnf.append(lits)                # at-least-one
        add_at_most_one(cnf, lits)      # at-most-one

    # NEW: per-slot AMO — đảm bảo một y chỉ do tối đa 1 x “kích hoạt”
    for (_, _, _), lits in bucket.items():
        add_at_most_one(cnf, lits)    

    # build conflict graph among y on each machine
    def overlap(a1, a2, b1, b2):
        # [a1,a2) overlaps [b1,b2)
        return (a1 < b2) and (b1 < a2)

    y_nodes_per_m = {}
    for (m, t1, t2, yv) in active_y_nodes:
        y_nodes_per_m.setdefault(m, []).append((t1, t2, yv))

    # biclique cover encoding for at-most-one among pairwise-overlapping y
    for m in range(M):
        nodes = y_nodes_per_m.get(m, [])
        if not nodes:
            continue

        idx_to_node = {idx: nodes[idx] for idx in range(len(nodes))}
        edges = set()
        for u in range(len(nodes)):
            t1u, t2u, _ = idx_to_node[u]
            for v in range(u + 1, len(nodes)):
                t1v, t2v, _ = idx_to_node[v]
                if overlap(t1u, t2u, t1v, t2v):
                    edges.add((u, v))

        if not edges:
            continue

        neighbors = {i: set() for i in range(len(nodes))}
        for (u, v) in edges:
            neighbors[u].add(v)
            neighbors[v].add(u)

        def has_edge(a, b):
            if a == b: return False
            u, v = (a, b) if a < b else (b, a)
            return (u, v) in edges

        # simple greedy biclique cover that DOES NOT drop edges before they’re covered
        uncovered = set(edges)
        while uncovered:
            # pick a seed edge
            u0, v0 = next(iter(uncovered))
            U = {u0}
            V = {v0}

            changed = True
            while changed:
                changed = False
                # try to grow U: a node w can join U if connected to all V
                for w in range(len(nodes)):
                    if w in U or w in V: continue
                    if all(has_edge(w, vv) for vv in V):
                        U.add(w); changed = True
                # try to grow V: a node w can join V if connected from all U
                for w in range(len(nodes)):
                    if w in U or w in V: continue
                    if all(has_edge(uu, w) for uu in U):
                        V.add(w)
                        changed = True

            # add aux var and clauses for this biclique
            aux = cnf.nv + 1
            cnf.nv = aux
            for u in sorted(U):
                _, _, y_u = idx_to_node[u]
                cnf.append([-y_u, aux])      # y_U -> aux
            for v in sorted(V):
                _, _, y_v = idx_to_node[v]
                cnf.append([-aux, -y_v])     # aux -> ¬y_V
            # cnf.append([-aux] + [-y for (_,_,y) in [idx_to_node[u] for u in U+V]])

            # mark covered edges (all pairs U×V)
            covered = set()
            for uu in U:
                for vv in V:
                    a, b = (uu, vv) if uu < vv else (vv, uu)
                    if (a, b) in uncovered:
                        covered.add((a, b))
            uncovered.difference_update(covered)

    return cnf

# ----- solver harness -----
def solve_and_record(task_id, problem_name, N, M, T, d, r, e):
    start = time.time()
    cnf = encode_schedule_problem(N, M, T, d, r, e)

    solver = Cadical()
    for clause in cnf.clauses:
        solver.add_clause(clause)

    status = solver.solve()
    elapsed = time.time() - start

    if status:
        model = set(solver.get_model())
        log("Solution is valid!")
        for i in range(N):
            dur = d[i]
            assigned = False
            t_start = max(0, r[i])
            t_end_allowed = min(e[i] - dur, T - dur)
            for t0 in range(t_start, t_end_allowed + 1):
                for m in range(M):
                    if var_x(i, t0, m, N, M, T) in model:
                        log(f"Task {i} -> machine {m} : [{t0}, {t0 + dur})")
                        assigned = True
                        break
                if assigned: break
            if not assigned:
                log(f"Task {i}: (no start in model)")
    else:
        log("UNSAT")

    result = {
        "ID": task_id,
        "Problem": problem_name,
        "Type": "biclique",
        "Time": round(elapsed, 4),
        "Result": "SAT" if status else "UNSAT",
        "Variables": solver.nof_vars(),
        "Clauses": solver.nof_clauses()
    }
    solver.delete()
    write_to_excel(result)

def process_input_dir(input_dir, resource=20):
    task_id = 1
    for filename in sorted(os.listdir(input_dir)):
        if not filename.endswith(".txt"): continue
        path = os.path.join(input_dir, filename)
        with open(path) as f:
            # line 1: N
            N = int(f.readline().strip())
            # line 2: [(release, duration, deadline), ...]
            tasks = ast.literal_eval(f.readline().strip())
            r = [task[0] for task in tasks]
            d = [task[1] for task in tasks]
            e = [task[2] for task in tasks]

            # resource = len(tasks)
            T = max(e) if e else 0

            log(f"\n Processing: {filename} | N={N} M={resource} T={T}")
            solve_and_record(task_id, filename, N, resource, T, d, r, e)
            task_id += 1
    log_file.close()

if __name__ == "__main__":
    input_dir = os.path.join("input", sys.argv[1])
    process_input_dir(input_dir)
