import ast
import os
import time
import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from pysat.solvers import Glucose3
from pysat.formula import CNF

M = 4  # Số máy (có thể điều chỉnh)
# ---------- Log ----------
log_file = open('run.log', 'a')
def log(*args, **kwargs):
    print(*args, **kwargs)
    print(*args, file=log_file, **kwargs)
    log_file.flush()

# ---------- Ghi Excel ----------
def write_to_excel(result_dict):
    df = pd.DataFrame([result_dict])
    date = datetime.now().strftime('%Y-%m-%d')
    output_dir = 'out'
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    file_path = f'{output_dir}/results_{date}.xlsx'

    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
        except BadZipFile:
            book = Workbook()
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')
        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(file_path)
    else:
        df.to_excel(file_path, index=False, sheet_name='Results', header=True)

    log(f"Ghi kết quả vào Excel: {os.path.abspath(file_path)}")

# ---------- Hàm biến ----------
def var_x(i, t, m, M, T):
    return 1 + i * M * T + t * M + m

def y_offset(N, M, T):
    return 1 + N * M * T

def var_y(m, t1, t2, M, T, N):
    return y_offset(N, M, T) + m * T * T + t1 * T + t2

# ---------- Mã hóa bài toán ----------
def encode_schedule_problem(N, M, T, d, r, e):
    cnf = CNF()
    active_y = set()

    """ R1:  x_{i,t,m} -> y_{m,t,t+d[i]} 
    Nếu tác vụ i bắt đầu tại t trên máy m sẽ chiếm máy m từ t đến t+d[i]
    """
    for i in range(N):
        for m in range(M):
            for t in range(r[i], e[i] - d[i] + 1):
                t_end = t + d[i]
                xi = var_x(i, t, m, M, T)
                yi = var_y(m, t, t_end, M, T, N)
                cnf.append([-xi, yi])
                active_y.add((m, t, t_end, yi))

    """ R2: 
    Mỗi tác vụ i, tồn tại ít nhất một cặp (m, t) sao cho x_{i,t,m} đúng
    """
    for i in range(N):
        lits = []
        for m in range(M):
            for t in range(r[i], e[i] - d[i] + 1):
                lits.append(var_x(i, t, m, M, T))
        
        if lits:
            cnf.append(lits)
            for a in range(len(lits)):
                for b in range(a + 1, len(lits)):
                    cnf.append([-lits[a], -lits[b]])

    # # R3: các y không chồng lấn (trên cùng máy)
    # for m in range(M):
    #     intervals = [(t1, t2) for (mach, t1, t2) in active_y if mach == m]
    #     for i in range(len(intervals)):
    #         for j in range(i + 1, len(intervals)):
    #             t1, t2 = intervals[i]
    #             t3, t4 = intervals[j]
    #             if not (t2 <= t3 or t4 <= t1):  # chồng lấn
    #                 y1 = var_y(m, t1, t2, M, T, N)
    #                 y2 = var_y(m, t3, t4, M, T, N)
    #                 cnf.append([-y1, -y2])
    # R3: các y không chồng lấn (trên cùng máy) theo AMO_PE
    for m in range(M):
        intervals = [(t1, t2, yi) for (mach, t1, t2, yi) in active_y if mach == m]
        for i in range(len(intervals)):
            t1, t2, y1 = intervals[i]
            for j in range(i + 1, len(intervals)):
                t3, t4, y2 = intervals[j]
                if t1 < t4 and t3 < t2:  # Kiểm tra giao nhau chính xác
                    cnf.append([-y1, -y2])

    return cnf

# ---------- Giải và ghi kết quả ----------
def solve_and_record(task_id, problem_name, N, M, T, d, r, e):
    log(f"\n🔍 Bắt đầu giải bài toán: {problem_name} (ID: {task_id})")
    log(f"Task({r}, {e}, {d}), Machines: {M}, Time slots: {T}")
    start = time.time()
    cnf = encode_schedule_problem(N, M, T, d, r, e)
    solver = Glucose3()
    for clause in cnf.clauses:
        solver.add_clause(clause)

    status = solver.solve()
    elapsed = time.time() - start

    if status:
        model = solver.get_model()
        log("→ Lịch tìm được:")
        for i in range(N):
            for t in range(T):
                for m in range(M):
                    if var_x(i, t, m, M, T) in model:
                        log(f"Tác vụ {i} gán cho máy {m} từ thời điểm {t} đến {t + e[i]}")
                        break
                else:
                    continue
                break
    else:
        log("⛔ Không tìm được lịch hợp lệ.")

    result = {
        "ID": task_id,
        "Problem": problem_name,
        "Type": "biclique",
        "Time": round(elapsed, 4),
        "Result": "SAT" if status else "UNSAT",
        "Variables": solver.nof_vars(),
        "Clauses": solver.nof_clauses()
    }
    solver.delete()
    write_to_excel(result)

# ---------- Xử lý thư mục đầu vào ----------
def process_input_dir(input_dir):
    task_id = 1
    for filename in sorted(os.listdir(input_dir)):
        if filename.endswith(".txt"):
            path = os.path.join(input_dir, filename)
            with open(path) as f:
                N = int(f.readline().strip())
                tasks = ast.literal_eval(f.readline().strip())
                r = [task[0] for task in tasks]
                d = [task[1] for task in tasks]
                e = [task[2] for task in tasks]
                T = max(e)  # Thời gian tối đa cần xét
                log(f"\n🎯 Đang xử lý: {filename}")
                solve_and_record(task_id, filename, N, M, T, d, r, e)
                task_id += 1
    log_file.close()

# ---------- Gọi từ terminal ----------
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("❗ Sử dụng: python scheduler_sat.py <tên_thư_mục_input>")
    else:
        input_dir = os.path.join("input", sys.argv[1])
        process_input_dir(input_dir)